# coding:utf-8
import calendar
import datetime
import pandas as pd 
#import os
import random
from openpyxl import load_workbook
from openpyxl.styles import Border,Side
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class schedule(object):
	def __init__(self):
		self.weightDays = []
		self.numOfMember = 0
		self.membersWorkDay = []
		self.memberNotWorkDay = []
		self.members = []
		self.membersScore = []
		self.membersRank = []
		self.workTable = []
		self.startDay = 0
		self.membersLastWorkDay = []
		self.firstDayIsSatNeedSchedule = False

	def getWeightDays(self):
		today = datetime.datetime.today()
		month = today.month
		year = today.year
		c = calendar
		# 0 for monday, 1 for tuesday
		startDay, totalDays= c.monthrange(year, month+1)
		# let 1 be monday
		startDay += 1
		#print(startDay, totalDays)
		tmpDay = startDay
		self.startDay = startDay
		if self.startDay == 6:
			self.firstDayIsSatNeedSchedule = True
		weightDays = [0]*totalDays
		#print(len(weightDays))
		for i in range(len(weightDays)):
			if tmpDay >= 6:
				weightDays[i] = 2
			else:
				weightDays[i] = 1
			tmpDay = (tmpDay%7)+1
		self.weightDays = weightDays
		#print(self.weightDays)

	def readHolidays(self, fileName, sheetName):
		df = pd.read_excel(fileName,engine='openpyxl', sheet_name=sheetName)
		dfList = list(df)
		for i, row in enumerate(df):
			for j, c in enumerate(df[row]):
				self.weightDays[c-1] = 2
		#print(self.weightDays)

	def readSheet(self, fileName, sheetName):
		df = pd.read_excel(fileName,engine='openpyxl', sheet_name=sheetName)
		#print(type(df))
		dfList = list(df)
		self.numOfMember = len(dfList)
		#self.memberNotWorkDay = [self.numOfMember * [0]]
		self.memberNotWorkDay = [[] for i in range(self.numOfMember)]
		for i, row in enumerate(df):
			for j, c in enumerate(df[row]):
				if j == 0:
					self.membersRank.append(int(c))
				elif str(c) != "nan":
					#print(c, row, i)
					if c == "first half":
						for k in range(16):
							self.memberNotWorkDay[i].append(k)
					elif c == "second half":
						for k in range(16, 32):
							self.memberNotWorkDay[i].append(k)
					else:
						self.memberNotWorkDay[i].append(int(c))
				#print(str(c), type(c))
				
		#dfList = list(df[0:])
		self.members = list(df[0:])
		#print(self.memberNotWorkDay)

	def calculate(self):
		memberNotWorkIdx = [0 for i in range(self.numOfMember)]
		self.membersWorkDay = [[] for i in range(self.numOfMember)]
		self.membersScore = [0 for i in range(self.numOfMember)]
		self.workTable = [[] for i in range(len(self.weightDays))]
		self.membersLastWorkDay = [0 for i in range(self.numOfMember)]
		tmpDay = self.startDay
		for i, v in enumerate(self.weightDays):
			# select U1
			minScoreMember = 0
			minScore = self.membersScore[0]
			minScoreCandidate = []
			for j, u in enumerate(self.membersScore):
				choose = True
				for k, w in enumerate(self.memberNotWorkDay[j]):
					if (i+1) == w:
						choose = False
				if choose and self.membersRank[j] != 5 and \
					(self.membersLastWorkDay[j] == 0 or ((i+1) - self.membersLastWorkDay[j]) >= 3):
					minScoreCandidate.append(j)
			
			if minScoreCandidate != []:
				sortScore = []
				for j, u in enumerate(minScoreCandidate):
					sortScore.append(u)
				for j, u in enumerate(minScoreCandidate):
					tmpMin = self.membersScore[u]
					tmpMinIdx = j
					for k in range(j, len(minScoreCandidate)):
						if tmpMin > self.membersScore[minScoreCandidate[k]]:
							tmpMin = self.membersScore[minScoreCandidate[k]]
							#tmpMinIdx = minScoreCandidate[k]
							tmpMinIdx = k
					sortScore[j], sortScore[tmpMinIdx] = sortScore[tmpMinIdx], sortScore[j]
					#sortScore.append(tmpMinIdx)
				#print(sortScore)
				tmpMinScoreIdx = sortScore[0]
				#print(sortScore[0])
				#print(self.membersScore[tmpMinScoreIdx])
				randUpBound = -1
				selected = -1
				for j, u in enumerate(sortScore):
					#print(self.membersScore[u], self.membersScore[tmpMinScoreIdx])
					if self.membersScore[u] == self.membersScore[tmpMinScoreIdx]:
						randUpBound += 1
						# U select R3 first
						if self.membersRank[u] == 3:
							selected = j
				
				#print(minScoreCandidate)
				#print(randUpBound)
				if selected == -1:
					selected = random.randint(0,randUpBound)
				minScoreMember = sortScore[selected]
				#print(i+1, minScoreMember)
			else:
				sortScore = []
				for j, u in enumerate(self.membersScore):
					tmpMax = u
					tmpMaxIdx = j
					for k in range(j, len(self.membersScore)):
						if tmpMax < self.membersScore[k]:
							tmpMax = self.membersScore[k]
							tmpMaxIdx = k
					sortScore.append(tmpMaxIdx)
					if (self.membersRank[tmpMaxIdx] != 5) and (((i+1) - self.membersLastWorkDay[tmpMaxIdx]) >= 3):
						tmpSelect = True
						for k, w in enumerate(self.memberNotWorkDay[tmpMaxIdx]):
							if (i+1) == w:
								tmpSelect = False
						if tmpSelect:
							minScoreMember = tmpMaxIdx
			#print(minScoreMember)
			# selected
			self.membersLastWorkDay[minScoreMember] = i+1
			self.membersWorkDay[minScoreMember].append(i+1)
			
			if self.membersRank[minScoreMember] != 3:
				self.membersScore[minScoreMember] += v
			else:
				'''
				# R3 weekend or holiday points
				if v == 2:
					self.membersScore[minScoreMember] += 1
				else:
					self.membersScore[minScoreMember] += 0.7
				'''
				self.membersScore[minScoreMember] += (v*0.7)
			
			#self.membersScore[minScoreMember] += v
			self.workTable[i].append(minScoreMember)
			#print(minScoreMember)
			#print(self.membersScore)

			# select U2
			if tmpDay == 5:
				# same U2 at day 5 & 6
				minScoreMember = 0
				minScore = self.membersScore[0]
				minScoreCandidate = []
				for j, u in enumerate(self.membersScore):
					#if u == minScore:
					choose = True
					for k, w in enumerate(self.memberNotWorkDay[j]):
						if ((i+1) == w) or ((i+2) == w):
							choose = False
					if choose and self.membersRank[j] != 3 and \
						(self.membersLastWorkDay[j] == 0 or ((i+1) - self.membersLastWorkDay[j]) >= 3):
						minScoreCandidate.append(j)

				if minScoreCandidate != []:
					sortScore = []
					for j, u in enumerate(minScoreCandidate):
						sortScore.append(u)
					for j, u in enumerate(minScoreCandidate):
						tmpMin = self.membersScore[u]
						tmpMinIdx = j
						for k in range(j, len(minScoreCandidate)):
							if tmpMin > self.membersScore[minScoreCandidate[k]]:
								tmpMin = self.membersScore[minScoreCandidate[k]]
								#tmpMinIdx = minScoreCandidate[k]
								tmpMinIdx = k
						sortScore[j], sortScore[tmpMinIdx] = sortScore[tmpMinIdx], sortScore[j]
						#sortScore.append(tmpMinIdx)
					#print(sortScore)
					tmpMinScoreIdx = sortScore[0]
					randUpBound = -1
					for j, u in enumerate(sortScore):
						if self.membersScore[u] == self.membersScore[tmpMinScoreIdx]:
							randUpBound += 1
					
					#print(minScoreCandidate)
					#print(randUpBound)
					selected = random.randint(0,randUpBound)
					minScoreMember = sortScore[selected]
					#print(i+1, minScoreMember)
				else:
					sortScore = []
					for j, u in enumerate(self.membersScore):
						tmpMax = u
						tmpMaxIdx = j
						for k in range(j, len(self.membersScore)):
							if tmpMax < self.membersScore[k]:
								tmpMax = self.membersScore[k]
								tmpMaxIdx = k
						sortScore.append(tmpMaxIdx)
						if (self.membersRank[tmpMaxIdx] != 3) and (((i+1) - self.membersLastWorkDay[tmpMaxIdx]) >= 3):
							tmpSelect = True
							for k, w in enumerate(self.memberNotWorkDay[tmpMaxIdx]):
								if (i+1) == w:
									tmpSelect = False
							if tmpSelect:
								minScoreMember = tmpMaxIdx
				# selected
				#self.membersLastWorkDay[minScoreMember] = i+2
				self.membersLastWorkDay[minScoreMember] = i+1
				self.membersWorkDay[minScoreMember].append(i+1)
				#self.membersWorkDay[minScoreMember].append(i+2)
				#self.membersScore[minScoreMember] += (v + 2)
				self.membersScore[minScoreMember] += v
				self.workTable[i].append(minScoreMember)
				self.workTable[i+1].append(minScoreMember)
				#print(minScoreMember)
				#print(self.membersScore)
			elif tmpDay == 6:
				if self.firstDayIsSatNeedSchedule:
					self.firstDayIsSatNeedSchedule = False
					minScoreMember = 0
					minScore = self.membersScore[0]
					minScoreCandidate = []
					for j, u in enumerate(self.membersScore):
						#if u == minScore:
						choose = True
						for k, w in enumerate(self.memberNotWorkDay[j]):
							if ((i+1) == w) or ((i+2) == w):
								choose = False
						if choose and self.membersRank[j] != 3 and \
							(self.membersLastWorkDay[j] == 0 or ((i+1) - self.membersLastWorkDay[j]) >= 3):
							minScoreCandidate.append(j)

					if minScoreCandidate != []:
						sortScore = []
						for j, u in enumerate(minScoreCandidate):
							sortScore.append(u)
						for j, u in enumerate(minScoreCandidate):
							tmpMin = self.membersScore[u]
							tmpMinIdx = j
							for k in range(j, len(minScoreCandidate)):
								if tmpMin > self.membersScore[minScoreCandidate[k]]:
									tmpMin = self.membersScore[minScoreCandidate[k]]
									#tmpMinIdx = minScoreCandidate[k]
									tmpMinIdx = k
							sortScore[j], sortScore[tmpMinIdx] = sortScore[tmpMinIdx], sortScore[j]
							#sortScore.append(tmpMinIdx)
						#print(sortScore)
						tmpMinScoreIdx = sortScore[0]
						randUpBound = -1
						for j, u in enumerate(sortScore):
							if self.membersScore[u] == self.membersScore[tmpMinScoreIdx]:
								randUpBound += 1
						
						#print(minScoreCandidate)
						#print(randUpBound)
						selected = random.randint(0,randUpBound)
						minScoreMember = sortScore[selected]
						#print(i+1, minScoreMember)
					else:
						sortScore = []
						for j, u in enumerate(self.membersScore):
							tmpMax = u
							tmpMaxIdx = j
							for k in range(j, len(self.membersScore)):
								if tmpMax < self.membersScore[k]:
									tmpMax = self.membersScore[k]
									tmpMaxIdx = k
							sortScore.append(tmpMaxIdx)
							if (self.membersRank[tmpMaxIdx] != 3) and (((i+1) - self.membersLastWorkDay[tmpMaxIdx]) >= 3):
								tmpSelect = True
								for k, w in enumerate(self.memberNotWorkDay[tmpMaxIdx]):
									if (i+1) == w:
										tmpSelect = False
								if tmpSelect:
									minScoreMember = tmpMaxIdx
					# selected
					self.membersLastWorkDay[minScoreMember] = i+2
					self.membersWorkDay[minScoreMember].append(i+1)
					self.membersWorkDay[minScoreMember].append(i+2)
					#self.membersScore[minScoreMember] += (v + 2)
					self.membersScore[minScoreMember] += v
					self.workTable[i].append(minScoreMember)
					self.workTable[i+1].append(minScoreMember)
					#print(minScoreMember)
					#print(self.membersScore)
			else:
				minScoreMember = 0
				minScore = self.membersScore[0]
				minScoreCandidate = []
				for j, u in enumerate(self.membersScore):
					#if u == minScore:
					choose = True
					for k, w in enumerate(self.memberNotWorkDay[j]):
						if (i+1) == w:
							choose = False
					if choose and self.membersRank[j] != 3 and \
						(self.membersLastWorkDay[j] == 0 or ((i+1) - self.membersLastWorkDay[j]) >= 3):
						minScoreCandidate.append(j)

				if minScoreCandidate != []:
					sortScore = []
					for j, u in enumerate(minScoreCandidate):
						sortScore.append(u)
					for j, u in enumerate(minScoreCandidate):
						tmpMin = self.membersScore[u]
						tmpMinIdx = j
						for k in range(j, len(minScoreCandidate)):
							if tmpMin > self.membersScore[minScoreCandidate[k]]:
								tmpMin = self.membersScore[minScoreCandidate[k]]
								#tmpMinIdx = minScoreCandidate[k]
								tmpMinIdx = k
						sortScore[j], sortScore[tmpMinIdx] = sortScore[tmpMinIdx], sortScore[j]
						#sortScore.append(tmpMinIdx)
					#print(sortScore)
					tmpMinScoreIdx = sortScore[0]
					randUpBound = -1
					for j, u in enumerate(sortScore):
						
						#print(self.membersScore[u], self.membersScore[tmpMinScoreIdx])
						if self.membersScore[u] == self.membersScore[tmpMinScoreIdx]:
							randUpBound += 1
					
					#print(minScoreCandidate)
					#print(randUpBound)
					selected = random.randint(0,randUpBound)
					minScoreMember = sortScore[selected]
					#print(i+1, minScoreMember)
				else:
					sortScore = []
					for j, u in enumerate(self.membersScore):
						tmpMax = u
						tmpMaxIdx = j
						for k in range(j, len(self.membersScore)):
							if tmpMax < self.membersScore[k]:
								tmpMax = self.membersScore[k]
								tmpMaxIdx = k
						sortScore.append(tmpMaxIdx)
						if (self.membersRank[tmpMaxIdx] != 3) and (((i+1) - self.membersLastWorkDay[tmpMaxIdx]) >= 3):
							tmpSelect = True
							for k, w in enumerate(self.memberNotWorkDay[tmpMaxIdx]):
								if (i+1) == w:
									tmpSelect = False
							if tmpSelect:
								minScoreMember = tmpMaxIdx
				# selected
				self.membersLastWorkDay[minScoreMember] = i+1
				self.membersWorkDay[minScoreMember].append(i+1)
				self.membersScore[minScoreMember] += v
				self.workTable[i].append(minScoreMember)
				#print(minScoreMember)
				#print(self.membersScore)

			tmpDay = (tmpDay%7)+1
		print(self.workTable)
		print(self.membersScore)
		for i, v in enumerate(self.membersWorkDay):
			print(len(v))
		#print(self.membersWorkDay)

	def set_border(self, cell):
		'''
		rows = ws.range(cell_range)
		for row in rows:
			row[0].style.borders.left.border_style = Border.BORDER_THIN
			row[-1].style.borders.right.border_style = Border.BORDER_THIN
		for c in rows[0]:
			c.style.borders.top.border_style = Border.BORDER_THIN
		for c in rows[-1]:
			c.style.borders.bottom.border_style = Border.BORDER_THIN
		'''
		border = Border(left=Side(border_style='thin',color='000000'),
		right=Side(border_style='thin',color='000000'),
		top=Side(border_style='thin',color='000000'),
		bottom=Side(border_style='thin',color='000000'))
		cell.border = border

	def writeSheet(self):
		# green
		fill = PatternFill("solid", fgColor="00CCFFCC")
		wb = Workbook()
		ws =  wb.active
		ws.title = "Sheet"
		sheet = wb[ws.title]
		#self.set_border(sheet, "C3:H10")
		rowOffset = 1
		columnOffset = 1
		translateDay = ["零", "一", "二", "三", "四", "五", "六", "日"]
		# start from 1 not 0
		#sheet.cell(row=4, column=1, value="123")
		cell = sheet.cell(row=1, column=1)
		self.set_border(cell)
		cell = sheet.cell(row=1, column=2)
		self.set_border(cell)
		sheet.cell(row=1, column=1, value="日期")
		sheet.cell(row=1, column=2, value="星期")
		columnOffset += 2
		
		for i, v in enumerate(self.members):
			sheet.cell(row=1, column=i+3, value=v)
			cell = sheet.cell(row=1, column=columnOffset + i)
			self.set_border(cell)

		
		rowOffset += 1
		tmpStartDay = self.startDay
		for i, v in enumerate(self.weightDays):
			if v == 2:
				sheet.cell(row=rowOffset + i, column=1, value=i+1).fill = fill
				sheet.cell(row=rowOffset + i, column=2, value=translateDay[tmpStartDay]).fill = fill

				for j, u in enumerate(self.members):
					cell = sheet.cell(row=rowOffset + i, column=columnOffset + j)
					cell.fill = fill
					#self.set_border(cell)
			

			else:
				sheet.cell(row=rowOffset + i, column=1, value=i+1)
				sheet.cell(row=rowOffset + i, column=2, value=translateDay[tmpStartDay])
			
			cell = sheet.cell(row=rowOffset + i, column=1)
			self.set_border(cell)
			cell = sheet.cell(row=rowOffset + i, column=2)
			self.set_border(cell)
			for j, u in enumerate(self.members):
				cell = sheet.cell(row=rowOffset + i, column=columnOffset + j)
				self.set_border(cell)
			
			tmpStartDay = (tmpStartDay%7)+1

		tmpStartDay = self.startDay
		fridayU = -1
		for i, v in enumerate(self.workTable):
			if tmpStartDay != 6:
				if self.membersRank[v[0]] < self.membersRank[v[1]]:
					sheet.cell(row=rowOffset + i, column=columnOffset + v[0], value='U')
					sheet.cell(row=rowOffset + i, column=columnOffset + v[1], value='U2')
					if tmpStartDay == 5:
						fridayU = v[0]
				else:
					sheet.cell(row=rowOffset + i, column=columnOffset + v[0], value='U2')
					sheet.cell(row=rowOffset + i, column=columnOffset + v[1], value='U')
					if tmpStartDay == 5:
						fridayU = v[1]
			else:
				if self.membersRank[v[0]] < self.membersRank[v[1]]:
					sheet.cell(row=rowOffset + i, column=columnOffset + v[0], value='U').fill = fill
					sheet.cell(row=rowOffset + i, column=columnOffset + v[1], value='▲').fill = fill
				else:
					sheet.cell(row=rowOffset + i, column=columnOffset + v[0], value='▲').fill = fill
					sheet.cell(row=rowOffset + i, column=columnOffset + v[1], value='U').fill = fill
				if fridayU != -1:
					sheet.cell(row=rowOffset + i, column=columnOffset + fridayU, value='△').fill = fill # △
					fridayU = -1
			tmpStartDay = (tmpStartDay%7)+1

		rowOffset += len(self.weightDays) + 1
		sheet.cell(row=rowOffset, column=2, value="Total點數")
		for i, v in enumerate(self.membersScore):
			sheet.cell(row=rowOffset, column=columnOffset + i, value=v)

		rowOffset += 1
		sheet.cell(row=rowOffset, column=2, value="Total排班數")
		for i, v in enumerate(self.membersWorkDay):
			sheet.cell(row=rowOffset, column=columnOffset + i, value=len(v))
		wb.save(filename = 'schedule.xlsx')
			

def main():
	#rand = random.randint(0,0) # 0, 0 is ok
	
	fileName = 'member.xlsx'
	scheduling = schedule()
	scheduling.getWeightDays()
	scheduling.readHolidays(fileName, 'Sheet1')
	scheduling.readSheet(fileName, 'Sheet2')
	scheduling.calculate()
	scheduling.writeSheet()
	

if __name__ == "__main__":
	main()